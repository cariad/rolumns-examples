from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from rolumns import Columns
from rolumns.renderers import RowsRenderer


def main() -> None:
    columns = Columns("people")
    columns.add("Name", "name")
    columns.add("Craft", "craft")

    data = requests.get("http://api.open-notify.org/astros.json").json()

    wb = Workbook()
    ws = wb.active

    for row in RowsRenderer(columns).render(data):
        ws.append(row)

    ws["A1"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 25

    ws["B1"].font = Font(bold=True)
    ws.column_dimensions["B"].width = 10

    path = (Path() / "who_in_space.xlsx").resolve().as_posix()

    wb.save(path)
    print(f"Exported to {path}")


if __name__ == "__main__":
    main()
